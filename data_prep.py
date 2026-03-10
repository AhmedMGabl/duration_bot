# -*- coding: utf-8 -*-
"""
data_prep.py
Reads rawdata.xlsx (sheets 1 & 2) and Team Structure.xlsx (CM sheet),
returns data ready for the dashboard /process API call.
"""
import io
import pandas as pd
import openpyxl
from openpyxl import Workbook


def load_crm_paste_text(rawdata_path: str) -> str:
    """
    Read CRM data from rawdata.xlsx sheet 1 (headers at row 0, no skiprows).
    Returns tab-separated text the dashboard paste parser can handle.
    """
    df = pd.read_excel(rawdata_path, sheet_name=0, engine='openpyxl')
    sc_col = df.columns[1]
    df = df[df[sc_col].notna()].copy()
    df = df[df[sc_col].astype(str).str.strip() != '']
    return df.to_csv(sep='\t', index=False)


def load_iur_paste_text(rawdata_path: str) -> str:
    """
    Read IUR data from rawdata.xlsx sheet 2.
    Dashboard parse_iur_data_from_text looks for:
      - First column: agent identifier
      - Column containing 'class completed'
      - Column containing 'attended students'
    Puts useraccount1 first, filters out group rows (小组).
    """
    df = pd.read_excel(rawdata_path, sheet_name=1, engine='openpyxl')
    agent_col = 'useraccount1' if 'useraccount1' in df.columns else df.columns[1]
    df = df[df[agent_col].notna()].copy()
    df = df[~df[agent_col].astype(str).str.contains('小组', na=False)]
    df = df[df[agent_col].astype(str).str.strip() != '']
    # Put agent column first so the dashboard uses it as the identifier
    cols = [agent_col] + [c for c in df.columns if c != agent_col]
    return df[cols].to_csv(sep='\t', index=False)


def load_cm_team_structure_bytes(team_structure_path: str) -> bytes:
    """
    Extract CM sheet from Team Structure.xlsx.
    Returns bytes of a single-sheet Excel (Team, CRM columns).
    Dashboard parse_team_structure_from_excel reads sheet 0, cols 0+1 as Team, CRM.
    """
    src_wb = openpyxl.load_workbook(team_structure_path)
    if 'CM' in src_wb.sheetnames:
        cm_ws = src_wb['CM']
    else:
        cm_ws = src_wb.active
        print(f"  WARNING: No 'CM' sheet found. Using first sheet: '{cm_ws.title}'")

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = 'CM'
    for row in cm_ws.iter_rows(values_only=True):
        if row[0] is not None or row[1] is not None:
            new_ws.append([row[0], row[1]])

    buf = io.BytesIO()
    new_wb.save(buf)
    buf.seek(0)
    return buf.read()


if __name__ == '__main__':
    import os
    base = os.path.dirname(__file__)
    rawdata = os.path.join(base, 'Input', 'rawdata.xlsx')
    structure = os.path.join(base, 'Input', 'Team Structure.xlsx')

    crm = load_crm_paste_text(rawdata)
    print(f"CRM rows: {len(crm.splitlines()) - 1}")
    print(f"CRM headers: {crm.splitlines()[0][:80]}")

    iur = load_iur_paste_text(rawdata)
    print(f"\nIUR rows: {len(iur.splitlines()) - 1}")
    print(f"IUR headers: {iur.splitlines()[0][:80]}")

    cm_bytes = load_cm_team_structure_bytes(structure)
    print(f"\nCM structure bytes: {len(cm_bytes)}")
